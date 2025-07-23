from django.shortcuts import render, redirect
from django.urls import reverse, reverse_lazy
from django.contrib.auth import login
from django.template import RequestContext

from django.core.mail import send_mail, BadHeaderError
from django.http import HttpResponse
from django.contrib import messages

from django.contrib.sites.shortcuts import get_current_site
from django.utils.encoding import force_bytes
from django.utils.encoding import force_str
from django.utils.http import urlsafe_base64_encode
from django.utils.http import urlsafe_base64_decode
from django.template.loader import render_to_string
from django.views.generic import View, UpdateView
from django.utils.translation import gettext as _
from .models import User
from django.http import HttpResponse
from myproj.firebase_config import initialize_firebase
import firebase_admin
import os, json
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from myproj.firebase_config import initialize_firebase
from myproj import settings

from django.http import JsonResponse # type:ignore
from django.contrib.auth.decorators import login_required # type:ignore
from pywebpush import webpush, WebPushException
import json
from datetime import datetime
from myproj import *
from django.contrib.auth.views import LoginView
from .models import FCMToken, Devices
from pyfcm import FCMNotification
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_protect, ensure_csrf_cookie
from django.middleware.csrf import get_token
from rest_framework import generics, viewsets, mixins, status
from .serializers import UserSerializer
from django.db.models import Q
from django.utils.dateparse import parse_date
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import inch
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.response import Response
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import login as auth_login
from django.contrib.auth import authenticate
from .permissions import IsSuperAdmin, IsAdmin
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny, IsAuthenticated
from .serializers import BackupSerializer


@method_decorator(ensure_csrf_cookie, name='dispatch')
@method_decorator(csrf_protect, name='dispatch')


class CustomLoginAPIView(APIView):
    """
    Vista de API para el login que:
    1. Proporciona CSRF token en GET
    2. Maneja autenticación en POST
    3. Envía notificaciones push al iniciar sesión
    """
    permission_classes = []  # Permitir acceso sin autenticación

    def get(self, request):
        """Endpoint para obtener CSRF token"""
        response = JsonResponse({'csrfToken': get_token(request)})
        response["Access-Control-Allow-Origin"] = request.headers.get('Origin', '*')
        response["Access-Control-Allow-Credentials"] = 'true'
        return response

    def post(self, request):
        """
        Maneja el login con:
        - Validación CSRF
        - Autenticación de usuario
        - Envío de notificaciones
        """
        # Verificación CSRF
        csrf_token = request.headers.get('X-CSRFToken')
        cookie_token = request.META.get('CSRF_COOKIE')
        
        if not csrf_token or not cookie_token or csrf_token != cookie_token:
            return Response({
                'status': 'error',
                'message': 'Token CSRF inválido'
            }, status=status.HTTP_403_FORBIDDEN)
        
        try:
            data = request.data
            user = authenticate(
                request,
                username=data.get('username'),
                password=data.get('password')
            )
            
            if user is not None:
            # Login exitoso
                login(request, user)
            
            # Configurar respuesta con el rol del usuario
                response = Response({
                    'status': 'success',
                    'user': {
                        'username': user.username,
                        'email': user.email,
                        'id': user.id,
                        'role': user.role  # Asegurarse de enviar el rol al frontend
                    }
                })
            
            # Establecer cookie con el rol del usuario
                response.set_cookie(
                    key='user_role',
                    value=user.role,
                    httponly=False,  # Para que Vue pueda leerla
                    secure=False,    # En producción debería ser True
                    samesite='Lax'
                )
            
            # Enviar notificaciones push
                self._send_login_notifications(user)
            
                return response
            
            return Response({
                'status': 'error',
                'message': 'Credenciales inválidas'
            }, status=status.HTTP_401_UNAUTHORIZED)
            
        except json.JSONDecodeError:
            return Response({
                'status': 'error',
                'message': 'Formato JSON inválido'
            }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'status': 'error',
                'message': f'Error en el servidor: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _send_login_notifications(self, user):
        """Envía notificaciones push sobre el inicio de sesión"""
        login_time = timezone.now().strftime("%Y-%m-%d %H:%M:%S")
        message = {
            "title": "Nuevo inicio de sesión",
            "body": f"El usuario {user.username} inició sesión el {login_time}",
            "icon": "/static/images/icon.png",
            "timestamp": login_time
        }
        
        # 1. Enviar notificación FCM (para apps móviles)
        self._send_fcm_notification(user, message)
        
        # 2. Enviar notificación Web Push (para navegadores)
        self._send_web_push_notification(user, message)

    def _send_fcm_notification(self, user, message):
        """Envía notificación a través de Firebase Cloud Messaging"""
        try:
            if hasattr(user, 'fcmtoken') and user.fcmtoken.is_active:
                push_service = FCMNotification(api_key=settings.FCM_API_KEY)
                result = push_service.notify_single_device(
                    registration_id=user.fcmtoken.token,
                    message_title=message["title"],
                    message_body=message["body"],
                    data_message={
                        "type": "login",
                        "timestamp": message["timestamp"],
                        "url": "/dashboard/"
                    }
                )
                print(f"Notificación FCM enviada a {user.username}")
        except Exception as e:
            print(f"Error en FCM para {user.username}: {str(e)}")

    def _send_web_push_notification(self, user, message):
        """Envía notificación Web Push a navegadores"""
        try:
            devices = Devices.objects.filter(user=user)
            for device in devices:
                try:
                    webpush(
                        subscription_info={
                            "endpoint": device.endpoint,
                            "keys": {
                                "p256dh": device.p256dh,
                                "auth": device.auth
                            }
                        },
                        data=json.dumps({
                            "notification": message,
                            "data": {
                                "type": "login",
                                "timestamp": message["timestamp"],
                                "url": "/dashboard/"
                            }
                        }),
                        vapid_private_key=settings.VAPID_PRIVATE_KEY,
                        vapid_claims={
                            "sub": f"mailto:{settings.CONTACT_EMAIL}",
                            "aud": "https://updates.push.services.mozilla.com"
                        }
                    )
                except WebPushException as e:
                    print(f"Error WebPush para {user.username}: {str(e)}")
                    if "410" in str(e):  # Gone status
                        device.delete()
        except Exception as e:
            print(f"Error general en WebPush para {user.username}: {str(e)}")
@ensure_csrf_cookie
def get_csrf(request):
    response = JsonResponse({'status': 'CSRF cookie set'})
    response["Access-Control-Allow-Origin"] = request.headers.get('Origin', '*')
    response["Access-Control-Allow-Credentials"] = 'true'
    return response

class DashboardDataAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        total_dispositivos = DispositivoInventario.objects.count()
        dispositivos_activos = DispositivoInventario.objects.filter(estado="Activo").count()
        total_usuarios = User.objects.count()
        
        chart_data = {
            'labels': ['Computadoras', 'Impresoras', 'Servidores', 'Redes', 'Otros'],
        }
        
        return Response({
            'total_dispositivos': total_dispositivos,
            'dispositivos_activos': dispositivos_activos,
            'total_usuarios': total_usuarios,
            'ultima_actualizacion': timezone.now().isoformat(),
        })

class UserProfileAPIView(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        return Response({
            'username': user.username,
            'full_name': user.get_full_name()
        })

  

class HomeAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        total_dispositivos = DispositivoInventario.objects.count()
        total_usuarios = User.objects.count()
        dispositivos_activos = DispositivoInventario.objects.filter(estado="Activo").count()
        
        return Response({
            'active_tab': 'home',
            'total_dispositivos': total_dispositivos,
            'total_usuarios': total_usuarios,
            'dispositivos_activos': dispositivos_activos
        })


class DispositivoInventarioAPIView(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        dispositivos = DispositivoInventario.objects.all()
        serializer = DispositivoInventarioSerializer(dispositivos, many=True)
        return Response({
            'dispositivos': serializer.data,
            'active_tab': 'inventario'
        })

    def post(self, request):
        serializer = DispositivoInventarioSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(creado_por=request.user)
            return Response({
                'status': 'success',
                'message': 'Dispositivo agregado correctamente',
                'data': serializer.data
            }, status=status.HTTP_201_CREATED)
        return Response({
            'status': 'error',
            'message': 'Error en los datos',
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import Group, Permission
from django.http import HttpResponse, HttpResponseForbidden
from django.contrib.auth.decorators import permission_required
User = get_user_model()

def is_superadmin_or_admin(user):
    return user.role in ['superadmin', 'admin'] or user.is_superuser


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsSuperAdmin]  # Solo superadmin puede manejar usuarios
    lookup_field = 'id'

    def perform_create(self, serializer):
        user = serializer.save()
        user.sync_role_to_group()
        # Ejemplo: Enviar email de bienvenida
        # send_welcome_email(user)

    @action(detail=True, methods=['post'])
    def activate(self, request, pk=None):
        user = self.get_object()
        user.is_active = True
        user.save()
        return Response({'status': 'user activated'})

    @action(detail=True, methods=['post'])
    def deactivate(self, request, pk=None):
        user = self.get_object()
        user.is_active = False
        user.save()
        return Response({'status': 'user deactivated'})

from rest_framework.permissions import IsAuthenticated
from .models import DispositivoInventario
from .serializers import DispositivoInventarioSerializer

class DispositivoInventarioViewSet(viewsets.ModelViewSet):
    """
    ViewSet para el manejo completo de CRUD de dispositivos de inventario.
    Se integra automáticamente con el router de DRF.
    """
    queryset = DispositivoInventario.objects.all()
    serializer_class = DispositivoInventarioSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['tipo', 'estado', 'ubicacion', 'responsable']

    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filtros personalizados
        fecha_inicio = self.request.query_params.get('fecha_inicio')
        fecha_fin = self.request.query_params.get('fecha_fin')
        
        if fecha_inicio and fecha_fin:
            queryset = queryset.filter(
                fecha_adquisicion__range=[fecha_inicio, fecha_fin]
            )
        
        return queryset

    def perform_create(self, serializer):
        serializer.save(creado_por=self.request.user)

    @action(detail=False, methods=['get'])
    def reporte(self, request):
        """
        Endpoint especial para generar reportes en diferentes formatos.
        Uso: /api/dispositivos/reporte/?formato=pdf|excel|json
        """
        dispositivos = self.filter_queryset(self.get_queryset())
        formato = request.query_params.get('formato', 'json')
        
        if formato == 'pdf':
            return generate_pdf_report(dispositivos)
        elif formato == 'excel':
            return generate_excel_report(dispositivos)
        
        serializer = self.get_serializer(dispositivos, many=True)
        return Response(serializer.data)

from django.shortcuts import render, redirect
from django.core.exceptions import ValidationError


@login_required
@permission_classes([IsAdmin])
def agregar_dispositivo(request):
    usuarios_disponibles = User.objects.all() 
    if request.method == 'POST':
        try:
            
            
            # Validación manual
            if not request.POST.get('serial_number'):
                raise ValidationError('El número de serie es requerido')
            
            responsable_id = request.POST.get('responsable')
            responsable = User.objects.get(id=responsable_id) if responsable_id else None


            dispositivo = DispositivoInventario(
                numero_inventario=request.POST['inventory_number'],
                tipo=request.POST['tipo'],
                marca=request.POST['brand'],
                modelo=request.POST['model'],
                serial=request.POST['serial_number'],
                ubicacion=request.POST['location'],
                fecha_adquisicion=request.POST['acquisition_date'],
                estado=request.POST['status'],
                responsable=responsable,
                creado_por=request.user
            )
            
            dispositivo.full_clean()  # Valida todos los campos
            dispositivo.save()
            
            messages.success(request, '✅ Dispositivo agregado correctamente!')
            return redirect('inventario')
            
        except Exception as e:
            messages.error(request, f'❌ Error: {str(e)}')
            print("Error al guardar:", str(e))  # Log para depuración
    
    return render(request, 'inventario.html', {'usuarios': usuarios_disponibles})
from django.core.paginator import Paginator
import logging
from openpyxl import Workbook
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from django.utils import timezone
logger = logging.getLogger(__name__)
@login_required
def lista_dispositivos(request):
    # 1. Verificar consulta a la base de datos
    dispositivos = DispositivoInventario.objects.all()
    
    
    # 2. Verificar contexto enviado al template
    context = {'dispositivos': dispositivos}
    print(f"DEBUG - Contexto enviado: {context}")
    
    return render(request, 'devices_list.html', context)
from .models import DispositivoInventario, Backup, User

from django.contrib.auth import logout
def custom_logout(request):
    logout(request)
    return redirect('login')


class ReportsBackupsViewSet(viewsets.ModelViewSet):
    queryset = Backup.objects.all().order_by('-fecha_creacion')
    serializer_class = BackupSerializer
    permission_classes = [IsAdmin]

    @action(detail=True, methods=['post'])
    def restore(self, request, pk=None):
        backup = self.get_object()
        # Aquí iría la lógica de restauración real
        return Response({
            'status': 'success',
            'message': f'Backup {backup.tipo} restaurado'
        })

    @action(detail=False, methods=['post'])
    def generate_report(self, request):
        dispositivos = DispositivoInventario.objects.all()
        
        # Aplicar filtros
        dispositivos = self._apply_filters(dispositivos, request.data)
        serializer = DispositivoInventarioSerializer(dispositivos, many=True)
        
        request.session['report_filters'] = request.data
        return Response(serializer.data)

    def _apply_filters(self, queryset, filters):
        # Implementación de filtros (similar a tu versión anterior)
        if filters.get('fecha_inicio') and filters.get('fecha_fin'):
            queryset = queryset.filter(
                fecha_adquisicion__range=[
                    filters['fecha_inicio'],
                    filters['fecha_fin']
                ]
            )
        if filters.get('estado'):
            queryset = queryset.filter(estado=filters['estado'])
        if filters.get('ubicacion'):
            queryset = queryset.filter(ubicacion=filters['ubicacion'])
        return queryset

from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from io import BytesIO

def generate_pdf_report(dispositivos):
    try:
        # Configurar el buffer y el documento
        buffer = BytesIO()
        
        # Usar SimpleDocTemplate para mejor control
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),  # Horizontal para tablas anchas
            rightMargin=inch/2,
            leftMargin=inch/2,
            topMargin=inch/2,
            bottomMargin=inch/2
        )
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        title_style.alignment = 1  # Centrado
        
        # Contenido del PDF
        elements = []
        
        # 1. Título
        elements.append(Paragraph("REPORTE DE INVENTARIO", title_style))
        elements.append(Spacer(1, 12))
        
        # 2. Tabla de datos
        data = [
            ["N° Inventario", "Tipo", "Marca", "Modelo", "Serial", "Ubicación", "Estado", "Responsable"]
        ]
        
        # Llenar datos
        for dispositivo in dispositivos:
            data.append([
                dispositivo.numero_inventario,
                dispositivo.get_tipo_display(),
                dispositivo.marca,
                dispositivo.modelo,
                dispositivo.serial,
                dispositivo.ubicacion,
                dispositivo.estado,
                dispositivo.responsable
            ])
        
        # Crear tabla con estilo
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2E7D32')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.beige),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 12))
        
        # 3. Pie de página
        elements.append(Paragraph(
            f"Total de dispositivos: {len(dispositivos)} | Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')}",
            styles['Normal']
        ))
        
        # Construir PDF
        doc.build(elements)
        
        # Configurar respuesta HTTP
        pdf = buffer.getvalue()
        buffer.close()
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_inventario.pdf"'
        response.write(pdf)
        
        return response
        
    except Exception as e:
        logger.error(f"Error generando PDF: {str(e)}")
        return HttpResponse(f"Error al generar PDF: {str(e)}", status=500)
    

def generate_excel_report(dispositivos):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"reporte_inventario_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Inventario"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Encabezados
    headers = ['N° Inventario', 'Tipo', 'Marca', 'Modelo', 'Serial', 'Ubicación', 'Estado', 'Responsable']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f'{col_letter}1']
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Datos
    for row_num, dispositivo in enumerate(dispositivos, 2):
        ws.cell(row=row_num, column=1, value=dispositivo.numero_inventario)
        ws.cell(row=row_num, column=2, value=dispositivo.get_tipo_display())
        ws.cell(row=row_num, column=3, value=dispositivo.marca)
        ws.cell(row=row_num, column=4, value=dispositivo.modelo)
        ws.cell(row=row_num, column=5, value=dispositivo.serial)
        ws.cell(row=row_num, column=6, value=dispositivo.ubicacion)
        ws.cell(row=row_num, column=7, value=dispositivo.estado)
        ws.cell(row=row_num, column=8, value=dispositivo.responsable)
        
        # Alternar colores de fila
        if row_num % 2 == 0:
            fill_color = 'F5F9F5'
        else:
            fill_color = 'FFFFFF'
        
        for col in range(1, 9):
            ws.cell(row=row_num, column=col).fill = PatternFill(
                start_color=fill_color, 
                end_color=fill_color, 
                fill_type='solid'
            )
    
    # Ajustar anchos de columna
    column_widths = [20, 15, 15, 25, 20, 25, 15, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Congelar encabezados
    ws.freeze_panes = 'A2'
    
    # Añadir información de generación
    ws.append([])
    ws.append(['Información del reporte:'])
    ws.append([f'Fecha de generación: {timezone.now().strftime("%d/%m/%Y %H:%M")}'])
    ws.append([f'Total de dispositivos: {len(dispositivos)}'])
    
    wb.save(response)
    return response

# Modificar export_report para usar filtros persistentes

class ExportReportAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, format):
        report_filters = request.session.get('report_filters', {})
        dispositivos = self._apply_filters(DispositivoInventario.objects.all(), report_filters)
        
        if format == 'pdf':
            return generate_pdf_report(dispositivos)
        elif format == 'excel':
            return generate_excel_report(dispositivos)
        
        return Response({
            'status': 'error',
            'message': 'Formato no soportado'
        }, status=status.HTTP_400_BAD_REQUEST)

    def _apply_filters(self, queryset, filters):
        # Reutilizar el mismo método de filtrado que en ReportsBackupsAPIView
        if filters.get('fecha_inicio') and filters.get('fecha_fin'):
            try:
                fecha_inicio_dt = parse_date(filters['fecha_inicio'])
                fecha_fin_dt = parse_date(filters['fecha_fin'])
                if fecha_inicio_dt and fecha_fin_dt:
                    queryset = queryset.filter(
                        fecha_adquisicion__range=[fecha_inicio_dt, fecha_fin_dt]
                    )
            except (ValueError, TypeError):
                pass
        
        if filters.get('estado'):
            queryset = queryset.filter(estado=filters['estado'])
        
        if filters.get('ubicacion'):
            queryset = queryset.filter(ubicacion=filters['ubicacion'])
        
        if filters.get('usuario_id'):
            try:
                usuario = User.objects.get(id=filters['usuario_id'])
                queryset = queryset.filter(responsable=usuario.username)
            except User.DoesNotExist:
                pass
        
        return queryset

# Vista para descargar backup (simulada)

class BackupOperationsAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, backup_id):
        try:
            backup = Backup.objects.get(id=backup_id)
            response = HttpResponse(content_type='application/octet-stream')
            response['Content-Disposition'] = f'attachment; filename="backup_{backup_id}.bak"'
            response.write(b"Simulated backup content")
            return response
        except Backup.DoesNotExist:
            return Response({
                'status': 'error',
                'message': 'Backup no encontrado'
            }, status=status.HTTP_404_NOT_FOUND)

    def post(self, request, backup_id):
        try:
            backup = Backup.objects.get(id=backup_id)
            return Response({
                'status': 'success',
                'message': f'Backup #{backup_id} restaurado exitosamente'
            })
        except Backup.DoesNotExist:
            return Response({
                'status': 'error',
                'message': 'Backup no encontrado'
            }, status=status.HTTP_404_NOT_FOUND)
from .models import Mantenimiento
from .serializers import MantenimientoSerializer

class MantenimientoViewSet(viewsets.ModelViewSet):
    queryset = Mantenimiento.objects.all().prefetch_related('dispositivo')
    serializer_class = MantenimientoSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # Filtros opcionales (ejemplo)
        queryset = super().get_queryset()
        dispositivo_id = self.request.query_params.get('dispositivo_id')
        if dispositivo_id:
            queryset = queryset.filter(dispositivo_id=dispositivo_id)
        return queryset

class FCMTokenAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            token = request.data.get('token')
            device_name = request.data.get('device_name', 'Unknown')

            if not token:
                return Response({
                    'success': False, 
                    'message': 'Token is required'
                }, status=status.HTTP_400_BAD_REQUEST)

            FCMToken.objects.update_or_create(
                user=request.user,
                defaults={
                    'token': token,
                    'is_active': True,
                    'device_name': device_name
                }
            )

            return Response({
                'success': True, 
                'message': 'Token saved successfully!'
            })
        except Exception as e:
            return Response({
                'success': False, 
                'message': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def get(self, request):
        has_token = FCMToken.objects.filter(
            user=request.user, 
            is_active=True
        ).exists()
        
        return Response({
            'hasToken': has_token
        })







def test_firestore_connection(db):
        try:
            doc_ref = db.collection("test_collection").document("test_doc")
            doc_ref.set({"message": "Hello from Django emulator!"})
            return True, "Successfully connected to Firestore emulator!"
        except Exception as e:
            return False, f"Error connecting to Firestore emulator: {e}"
        

def my_view(request):
    try:
        app, db, auth_client = initialize_firebase()

        if not test_firestore_connection(db):
            raise Exception("Failed to connect to Firestore emulator!") #Raise exception.

        # Load configuration from JSON file
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # Project root
        config_file_path = os.path.join(base_dir, '..', 'etc', 'config.json')

        with open(config_file_path, 'r') as f:
            config = json.load(f)

        user_email = config['FIREBASE_SECRETS']['user_email']
        user_password = config['FIREBASE_SECRETS']['user_password']

        # Example: Create a user in the Authentication emulator
        try:
            user = auth_client.create_user(
                email=user_email,
                password=user_password
            )
            print(f'Successfully created new user: {user.uid}')
        except Exception as e:
            print(f"Error creating user: {e}")

        # Example: Write data to Firestore emulator
        doc_ref = db.collection("users").document("alovelace")
        doc_ref.set({"first": "Ada", "last": "Lovelace", "born": 1815})

        return HttpResponse("Firebase interactions complete!")
    except FileNotFoundError as e:
        return HttpResponse(f"Config file error: {e}", status=500)
    except Exception as e:
        return HttpResponse(f"Error initializing Firebase: {e}", status=500)
    


class PushNotificationAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        notification_type = request.data.get('type')
        
        if notification_type == 'login':
            return self._handle_login_notification(request.user)
        elif notification_type == 'profile_visit':
            return self._handle_profile_visit(request)
        else:
            return Response({
                'status': 'error',
                'message': 'Tipo de notificación no válido'
            }, status=status.HTTP_400_BAD_REQUEST)

    def _handle_login_notification(self, user):
        login_time = timezone.now().strftime("%Y-%m-%d %H:%M:%S")
        message = {
            "title": "Nuevo inicio de sesión",
            "body": f"El usuario {user.username} inició sesión el {login_time}",
            "icon": "/static/images/icon.png",
            "timestamp": login_time
        }
        
        # Enviar notificaciones
        self._send_fcm_notification(user, message)
        self._send_web_push_notification(user, message)
        
        return Response({
            'status': 'success',
            'message': 'Notificaciones enviadas'
        })

    def _handle_profile_visit(self, request):
        username = request.data.get("username", request.user.username)
        visit_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        message = {
            "title": f"Visita de perfil",
            "body": f"{username} visitó el perfil el {visit_time}",
            "icon": "/static/images/icon.png"
        }
        
        self._send_web_push_notification(request.user, message)
        return Response({"status": "success"})

    def _send_fcm_notification(self, user, message):
        try:
            if hasattr(user, 'fcmtoken') and user.fcmtoken.is_active:
                push_service = FCMNotification(api_key=settings.FCM_API_KEY)
                result = push_service.notify_single_device(
                    registration_id=user.fcmtoken.token,
                    message_title=message["title"],
                    message_body=message["body"],
                    data_message={
                        "type": "login",
                        "timestamp": message["timestamp"],
                        "url": "/dashboard/"
                    }
                )
                return result
        except Exception as e:
            print(f"Error en FCM: {str(e)}")
            return None

    def _send_web_push_notification(self, user, message):
        try:
            devices = Devices.objects.filter(user=user)
            for device in devices:
                try:
                    webpush(
                        subscription_info={
                            "endpoint": device.endpoint,
                            "keys": {
                                "p256dh": device.p256dh,
                                "auth": device.auth
                            }
                        },
                        data=json.dumps({
                            "notification": message,
                            "data": {
                                "type": "login",
                                "timestamp": message.get("timestamp"),
                                "url": "/dashboard/"
                            }
                        }),
                        vapid_private_key=settings.VAPID_PRIVATE_KEY,
                        vapid_claims={
                            "sub": f"mailto:{settings.CONTACT_EMAIL}",
                            "aud": "https://updates.push.services.mozilla.com"
                        }
                    )
                except WebPushException as e:
                    if "410" in str(e):  # Gone status
                        device.delete()
        except Exception as e:
            print(f"Error en WebPush: {str(e)}")
def csrf_failure_view(request, reason=""):
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': False,
            'message': 'CSRF verification failed',
            'reason': reason
        }, status=403)
    # Fallback para solicitudes no-AJAX
    from django.views.csrf import csrf_failure
    return csrf_failure(request, reason=reason)

def save_token(request):
    if request.method == 'POST' and request.user.is_authenticated:
        try:
            data = json.loads(request.body.decode('utf-8'))
            token = data.get('token')

            if not token:
                return JsonResponse({'success': False, 'message': 'Token is required'}, status=400)

            # Actualizar o crear el token para el usuario actual
            FCMToken.objects.update_or_create(
                user=request.user,
                defaults={
                    'token': token,
                    'is_active': True,
                    'device_name': data.get('device_name', 'Unknown')
                }
            )

            return JsonResponse({'success': True, 'message': 'Token saved successfully!'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    return JsonResponse({'success': False, 'message': 'Authentication required'}, status=401)



@login_required
def check_fcm_token(request):
    try:
        has_token = FCMToken.objects.filter(user=request.user, is_active=True).exists()
        return JsonResponse({
            'status': 'success',
            'hasToken': has_token,
            'message': 'Token verificado'
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)
@login_required
def check_auth(request):
    return JsonResponse({'isAuthenticated': True})

@login_required
def check_fcm_token(request):
    has_token = FCMToken.objects.filter(user=request.user, is_active=True).exists()
    return JsonResponse({'hasToken': has_token})
@login_required      
def send_profile_visit_notification(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            username = data.get("username", request.user.username)

            from users.models import Device
            devices = Device.objects.filter(user=request.user)

            visit_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            message = {
                "title": f"Profile Visit Notification",
                "body": f"{username} visited profile a {visit_time}",
                "icon": "/static/images/icon.png"
            }

            for device in devices:
                try:
                    webpush(subscription_info={
                        "endpoint": device.endpoint,
                        "keys": {
                            "p265dh": device.p265dh,
                            "auth": device.auth
                        }
                    },
                    data= json.dumps(message),
                    vapid_private_key=settings.VAPID_PRIVATE_KEY,
                    vapid_claims={
                        "sub": f"mailto:{settings.FIREBASE_SECRETS.get('user_email', 'alextoledo2602@gmail.com')}"
                    })
                except WebPushException as ex:
                    print("Error al enviar notificacion", ex)
            return JsonResponse({"status": "Success"})
        except Exception as e:
            return JsonResponse({"status": "Error", "message": str(e)}, status=400)
    return JsonResponse({"status": "Error", "message": "Invalid request method"}, status=405)

class ProfileAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        return Response({
            'username': user.username,
            'email': user.email,
            'full_name': user.get_full_name()
        })
@login_required
def send_login_notification(request):
    if request.method == 'POST':
        try:
            user = request.user
            login_time = timezone.now().strftime("%Y-%m-%d %H:%M:%S")
            message = {
                "title": "Nuevo inicio de sesión",
                "body": f"El usuario {user.username} inició sesión el {login_time}",
                "icon": "/static/images/icon.png",
                "timestamp": login_time
            }
            print(f"Token FCM del usuario: {user.fcmtoken.token}")  # Debug
            
            push_service = FCMNotification(api_key=settings.FCM_API_KEY)
            response = push_service.notify_single_device(
                registration_id=user.fcmtoken.token,
                message_title="Prueba de notificación",
                message_body="Esta es una notificación de prueba",
                data_message={"url": "/dashboard/"}
            )
            
            print("Respuesta de FCM:", response)
            # Enviar notificación FCM
            if hasattr(user, 'fcmtoken') and user.fcmtoken.is_active:
                push_service = FCMNotification(api_key=settings.FCM_API_KEY)
                result = push_service.notify_single_device(
                    registration_id=user.fcmtoken.token,
                    message_title=message["title"],
                    message_body=message["body"],
                    data_message={
                        "type": "login",
                        "timestamp": message["timestamp"],
                        "url": "/accounts/dashboard/"
                    }
                )
                print("Resultado FCM:", result)
            
            # Enviar notificación Web Push
            devices = Devices.objects.filter(user=user)
            for device in devices:
                try:
                    webpush(
                        subscription_info={
                            "endpoint": device.endpoint,
                            "keys": {
                                "p256dh": device.p256dh,
                                "auth": device.auth
                            }
                        },
                        data=json.dumps({
                            "notification": message,
                            "data": {
                                "type": "login",
                                "timestamp": message["timestamp"],
                                "url": "/accounts/dashboard/"
                            }
                        }),
                        vapid_private_key=settings.VAPID_PRIVATE_KEY,
                        vapid_claims={
                            "sub": f"mailto:{settings.CONTACT_EMAIL}",
                            "aud": "https://updates.push.services.mozilla.com"
                        }
                    )
                except WebPushException as e:
                    print("Error WebPush:", e)
                    if "410" in str(e):
                        device.delete()
            
            return JsonResponse({"status": "success", "message": "Notificaciones enviadas"})
        
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)
    
    return JsonResponse({"status": "error", "message": "Método no permitido"}, status=405)
class UserListCreateView(generics.ListCreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer

class UserRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer
